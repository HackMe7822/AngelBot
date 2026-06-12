import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sqlite3
import os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from data.database import get_conn

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
BLUE   = PatternFill("solid", fgColor="1F497D")
LBLUE  = PatternFill("solid", fgColor="DEEAF1")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=14, color="1F497D")

def thin_border():
    side = Side(style='thin', color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def generate_daily_report(date_str=None):
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")

    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        SELECT symbol, entry_time, exit_time, entry_price, exit_price,
               quantity, capital_used, pnl, pnl_pct, stop_loss, target, exit_reason
        FROM trades WHERE status='closed' AND (source='paper' OR source IS NULL) AND date(exit_time)=?
        ORDER BY exit_time
    """, (date_str,))
    trades = c.fetchall()
    c.execute("SELECT SUM(pnl) FROM trades WHERE status='closed' AND (source='paper' OR source IS NULL)")
    total_pnl = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM top_ups")
    reload_count, total_reloaded = c.fetchone()
    c.execute("SELECT amount, balance_before, time FROM top_ups WHERE date(time)=? ORDER BY time", (date_str,))
    todays_reloads = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Trades {date_str}"

    # Title
    ws.merge_cells("A1:L1")
    ws['A1'] = f"AngelBot Daily Report — {date_str}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Summary row
    day_pnl        = sum(t[7] for t in trades)
    wins           = sum(1 for t in trades if t[7] > 0)
    losses         = len(trades) - wins
    total_invested = 1000 + (total_reloaded or 0)
    balance        = total_invested + total_pnl
    net_pnl        = balance - total_invested

    ws.merge_cells("A2:C2"); ws['A2'] = f"Total Trades: {len(trades)}"
    ws.merge_cells("D2:E2"); ws['D2'] = f"Won: {wins}  |  Lost: {losses}"
    ws.merge_cells("F2:G2"); ws['F2'] = f"Day P&L: ₹{day_pnl:+.2f}"
    ws.merge_cells("H2:I2"); ws['H2'] = f"Balance: ₹{balance:.2f}"
    ws.merge_cells("J2:L2"); ws['J2'] = f"Net P&L: ₹{net_pnl:+.2f}  |  Reloads: {reload_count or 0}×"
    for cell in ['A2','D2','F2','H2','J2']:
        ws[cell].font = Font(bold=True, size=11)
        ws[cell].fill = LBLUE

    ws.append([])

    # Headers
    headers = ["Symbol","Entry Time","Exit Time","Entry ₹","Exit ₹",
               "Qty","Capital ₹","P&L ₹","P&L %","Stop-Loss","Target","Exit Reason"]
    ws.append(headers)
    hr = ws.max_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col)
        cell.font = HEADER_FONT
        cell.fill = BLUE
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    # Data rows
    for i, t in enumerate(trades):
        row = list(t)
        ws.append(row)
        dr = ws.max_row
        fill = GREEN if t[7] >= 0 else RED
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=dr, column=col)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal='center')

    # Top-ups section (today's reloads)
    if todays_reloads:
        ws.append([])
        ws.append(["AUTO-RELOAD EVENTS TODAY"])
        tr = ws.max_row
        ws.cell(row=tr, column=1).font = Font(bold=True, size=11, color="FF6600")
        ORANGE = PatternFill("solid", fgColor="FCE4D6")
        reload_headers = ["Event", "Time", "Balance Before", "Amount Injected", "Balance After"]
        ws.append(reload_headers)
        rhr = ws.max_row
        for col, h in enumerate(reload_headers, 1):
            cell = ws.cell(row=rhr, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF6600")
            cell.alignment = Alignment(horizontal='center')
        for i, (amt, bal_before, t) in enumerate(todays_reloads, 1):
            ws.append([f"Reload #{i}", t, f"₹{bal_before:.2f}", f"₹{amt:.0f}", f"₹{bal_before + amt:.2f}"])
            dr = ws.max_row
            for col in range(1, 6):
                ws.cell(row=dr, column=col).fill = ORANGE
                ws.cell(row=dr, column=col).alignment = Alignment(horizontal='center')

    # Column widths
    widths = [12, 18, 18, 10, 10, 8, 10, 10, 8, 10, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    reports_dir = os.path.join(os.path.dirname(__file__), '..', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    path = os.path.join(reports_dir, f"trades_{date_str}.xlsx")
    wb.save(path)
    print(f"Report saved: {path}")
    return path

if __name__ == "__main__":
    path = generate_daily_report()
    print(f"Done: {path}")
